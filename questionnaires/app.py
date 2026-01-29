import streamlit as st
import pandas as pd
import random
import string
from openpyxl import Workbook, load_workbook
import os

st.set_page_config(page_title="アンケート管理", page_icon="📝")

EXCEL_FILE = "questionnaires.xlsx"

# --------------------
# ユーティリティ
# --------------------
def generate_id():
    return "".join(random.choices(string.ascii_lowercase + string.digits, k=15))

def valid_password(pw):
    return 1 <= len(pw) <= 15 and all(c in string.ascii_lowercase + string.digits for c in pw)

def get_wb():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "TOP"
        ws.append(["title", "id", "password", "one_time_only", "result_no_password"])
        wb.save(EXCEL_FILE)
    return load_workbook(EXCEL_FILE)

# --------------------
# URL解析
# --------------------
params = st.query_params

def normalize(v):
    if isinstance(v, list):
        return v[0]
    return v

page = normalize(params.get("page"))
qid = normalize(params.get("id"))

if page is None:
    page = "make_new"

# ====================
# 作成ページ
# ====================
if page == "make_new":
    st.title("アンケート新規作成")

    title = st.text_input("タイトル")
    password = st.text_input("編集・集計用パスワード", type="password")
    one_time = st.checkbox("1人1回のみ回答可", value=True)
    result_free = st.checkbox("集計ページをパスワードなしで公開")

    if st.button("作成"):
        if not title:
            st.error("タイトルを入力してください")
        elif not valid_password(password):
            st.error("パスワードは英小文字と数字のみ、1〜15文字です")
        else:
            wb = get_wb()
            ws = wb["TOP"]

            new_id = generate_id()
            ws.append([title, new_id, password, one_time, result_free])
            wb.create_sheet(new_id)
            wb.save(EXCEL_FILE)

            st.success("作成しました")
            st.code(f"?page=edit&id={new_id}")
            st.code(f"?page=answer&id={new_id}")

# ====================
# 回答ページ（①②③）
# ====================
elif page == "answer":
    if not qid:
        st.error("IDが指定されていません")
        st.stop()

    if f"answered_{qid}" not in st.session_state:
        st.session_state[f"answered_{qid}"] = False

    wb = get_wb()
    ws_top = wb["TOP"]
    record = None

    for row in ws_top.iter_rows(min_row=2, values_only=True):
        if row[1] == qid:
            record = row
            break

    if not record:
        st.error("アンケートが存在しません")
        st.stop()

    if record[3] and st.session_state[f"answered_{qid}"]:
        st.warning("このアンケートは1人1回までです")
        st.stop()

    ws = wb[qid]
    st.title(record[0])

    answers = []

    for col in range(2, ws.max_column + 1):
        q_type = ws.cell(row=1, column=col).value
        q_text = ws.cell(row=2, column=col).value

        if not q_type:
            continue

        st.markdown(q_text)

        if q_type == "1行記述":
            ans = st.text_input("", key=f"q{col}")
        elif q_type == "複数行記述":
            ans = st.text_area("", key=f"q{col}")
        elif q_type == "ラジオボタン":
            ans = st.radio("", ["はい", "いいえ"], key=f"q{col}")
        elif q_type == "チェックボックス":
            ans = st.checkbox("チェック", key=f"q{col}")
        elif q_type == "ドロップダウン":
            ans = st.selectbox("", ["選択してください", "A", "B", "C"], key=f"q{col}")
        elif q_type == "スライダー":
            ans = st.slider("", 0, 10, key=f"q{col}")
        else:
            ans = None

        answers.append(ans)

    if st.button("送信"):
        row = ws.max_row + 1
        for i, val in enumerate(answers):
            ws.cell(row=row, column=i + 2, value=str(val))
        wb.save(EXCEL_FILE)

        st.session_state[f"answered_{qid}"] = True
        st.success("回答ありがとうございました")

# ====================
# 編集ページ
# ====================
elif page == "edit":
    if not qid:
        st.error("IDが指定されていません")
        st.stop()

    wb = get_wb()
    ws_top = wb["TOP"]

    record = None
    for row in ws_top.iter_rows(min_row=2, values_only=True):
        if row[1] == qid:
            record = row
            break

    if not record:
        st.error("IDが存在しません")
        st.stop()

    st.title(f"編集ページ：{record[0]}")

    pw = st.text_input("パスワード", type="password")
    if pw != record[2]:
        st.warning("パスワードを入力してください")
        st.stop()

    ws = wb[qid]

    st.subheader("質問追加")

    q_type = st.selectbox(
        "質問タイプ",
        ["ラジオボタン", "ドロップダウン", "チェックボックス", "スライダー", "1行記述", "複数行記述"]
    )

    q_text = st.text_area("質問文（改行可・URLは自動リンク）")
    required = st.checkbox("必須")

    if st.button("質問を追加"):
        col = ws.max_column + 1 if ws.max_column >= 2 else 2
        ws.cell(row=1, column=col, value=q_type)
        ws.cell(row=2, column=col, value=f"{q_text}\n[必須]" if required else q_text)
        wb.save(EXCEL_FILE)
        st.success("質問を追加しました")

# ====================
# 結果ページ
# ====================
elif page == "result":
    if not qid:
        st.error("IDが指定されていません")
        st.stop()

    wb = get_wb()
    ws_top = wb["TOP"]

    record = None
    for row in ws_top.iter_rows(min_row=2, values_only=True):
        if row[1] == qid:
            record = row
            break

    if not record:
        st.error("IDが存在しません")
        st.stop()

    if not record[4]:
        pw = st.text_input("パスワード", type="password")
        if pw != record[2]:
            st.warning("パスワードが必要です")
            st.stop()

    ws = wb[qid]

    st.title(f"結果一覧：{record[0]}")

    headers = []
    data = []

    for col in range(2, ws.max_column + 1):
        headers.append(ws.cell(row=2, column=col).value)

    for row in range(3, ws.max_row + 1):
        data.append([ws.cell(row=row, column=col).value for col in range(2, ws.max_column + 1)])

    if data:
        df = pd.DataFrame(data, columns=headers)
        st.dataframe(df)
    else:
        st.info("まだ回答がありません")

# ====================
# 不正URL
# ====================
else:
    st.error("不正なページ指定です")
